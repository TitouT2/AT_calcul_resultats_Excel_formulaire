
import pandas as pd
import sqlite3
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, RadarChart, Reference

# Établir la connexion à la base de données SQLite3
conn = sqlite3.connect('AnalyseTransac.db')

# Créer un curseur pour exécuter les requêtes
cursor = conn.cursor()

# structure table result_egogramme (id_user INT, egogramme TEXT, valeur INT)
# Exécuter la requête SQL pour récupérer les données de la table "result_egogramme"
query = "SELECT egogramme,valeur FROM result_egogramme where id_user = 1"
cursor.execute(query)
rows = cursor.fetchall()

# Fermer la connexion à la base de données SQLite3
conn.close()
#creation du nouvel objet workbook
classeur = Workbook()

#creer la premiere feuille "Egogramme"
feuille_Egogramme = classeur.create_sheet(title="Egogramme")

#creer la premiere feuille "Egogramme"
feuille_drivers = classeur.create_sheet(title="Drivers")

#creer la deuxieme feuille "Postures"
feuille_postures = classeur.create_sheet(title="Postures")

# Écrire les en-têtes dans la feuille "egogramme"
feuille_Egogramme.append(["Egogramme", "valeur"])

# Écrire les en-têtes dans la feuille "egogramme"
feuille_drivers.append(["Drivers", "valeur"])

# Écrire les en-têtes dans la feuille "postures"
feuille_postures.append(["postures", "valeur"])

# Écrire les données dans les feuilles correspondantes
feuille_Egogramme.append(rows[0])
feuille_Egogramme.append(rows[1])
feuille_Egogramme.append(rows[2])
feuille_Egogramme.append(rows[3])
feuille_Egogramme.append(rows[4])
feuille_Egogramme.append(rows[5])
feuille_drivers.append(rows[6])
feuille_drivers.append(rows[7])
feuille_drivers.append(rows[8])
feuille_drivers.append(rows[9])
feuille_drivers.append(rows[10])
feuille_postures.append(rows[11])
feuille_postures.append(rows[12])
feuille_postures.append(rows[13])
feuille_postures.append(rows[14])
# Sauvegarder les modifications dans le classeur Excel
fichier_excel = "/Users/evelyne/Documents/ATresultquestionnaires/ResultAT_EVE.xlsx"
classeur.save(fichier_excel)

#faire le graphique Egogramme
#Accéder à la feuille "Egogramme"
feuille_Egogramme = classeur["Egogramme"]

# Récupérer les données des cellules A1:B6 (attention 7 ?)
donnees = []
for row in feuille_Egogramme.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2, values_only=True):
    donnees.append(row)


# Transposer les données pour préparer les valeurs x (egogramme) et y (valeur)
valeurs_x, valeurs_y = zip(*donnees)

# Créer le graphique à barres
graphique = BarChart()
graphique.title = "Graphique Egogramme"
graphique.x_axis.title = "Egogramme"
graphique.y_axis.title = "Valeur"

# Ajouter les données au graphique
donnees_valeurs = Reference(feuille_Egogramme, min_col=2, min_row=1, max_row=7)
donnees_egogramme = Reference(feuille_Egogramme, min_col=1, min_row=2, max_row=7)
graphique.add_data(donnees_valeurs, titles_from_data=True)
graphique.set_categories(donnees_egogramme)

# Ajouter le graphique à la feuille Excel
feuille_Egogramme.add_chart(graphique, "D1")

#-------------------
#feuille drivers
# Récupérer les données des cellules A1:B6
donnees = []
for row in feuille_drivers.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2, values_only=True):
    donnees.append(row)

# Transposer les données pour préparer les valeurs x (egogramme) et y (valeur)
valeurs_x, valeurs_y = zip(*donnees)

# Créer le graphique à barres
graphique = BarChart()
graphique.title = "Graphique Postures"
graphique.x_axis.title = "Postures"
graphique.y_axis.title = "Valeur"

# Ajouter les données au graphique
donnees_valeurs = Reference(feuille_drivers, min_col=2, min_row=1, max_row=6)
donnees_egogramme = Reference(feuille_drivers, min_col=1, min_row=2, max_row=6)
graphique.add_data(donnees_valeurs, titles_from_data=True)
graphique.set_categories(donnees_egogramme)

# Ajouter le graphique à la feuille Excel
feuille_drivers.add_chart(graphique, "D1")

#partie Postures
# Accéder à la feuille "Postures"
feuille_postures = classeur["Postures"]

# Récupérer les données des cellules A1:B5
donnees = []
for row in feuille_postures.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2, values_only=True):
    donnees.append(row)

# Transposer les données pour préparer les valeurs x (étiquettes) et y (valeurs)
etiquettes, valeurs = zip(*donnees)

# Créer le graphique radar
graphique = RadarChart()
graphique.title = "Graphique Radar Postures"
graphique.y_axis.majorGridlines = None

# Ajouter les données au graphique
donnees_valeurs = Reference(feuille_postures, min_col=2, min_row=1, max_row=5)
donnees_etiquettes = Reference(feuille_postures, min_col=1, min_row=2, max_row=5)
graphique.add_data(donnees_valeurs, titles_from_data=True)
graphique.set_categories(donnees_etiquettes)
graphique.style = 6
# Ajouter le graphique à la feuille Excel
feuille_postures.add_chart(graphique, "D1")


# Sauvegarder les modifications dans le classeur Excel
classeur.save(fichier_excel)

# N'oubliez pas de fermer le classeur une fois que vous avez terminé
classeur.close()



#attention charger Egogramme en premier et renommer ce que j'ai chargé en Drivers puis le mettre dans onglet Egogramme puis faire le graphique

